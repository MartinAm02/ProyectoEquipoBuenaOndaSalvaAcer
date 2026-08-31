"""Genera los xlsx Amazon_SO_WTD_* con el mismo formato de Vendor Central.

Reproduce byte a byte la estructura de la descarga manual: fila 1 con los
metadatos entre corchetes, fila 2 con los encabezados en espanol, y los
datos desde la fila 3. Los nulos de la API se escriben como celda vacia,
que es lo que hace Vendor Central.

    python export_xlsx.py --descarga 2026-08-30
    python export_xlsx.py --desde 2026-08-23 --hasta 2026-08-29
    python export_xlsx.py --descarga 2026-08-30 --destino "C:/ruta/donde/guardar"

OJO con la semana: Vendor Central usa la semana de AMAZON, domingo a
sabado. El archivo que se descarga el domingo 30 cubre del domingo 23 al
sabado 29. No es la semana ISO (lunes a domingo) que usa so_wk_XX.csv;
para esa esta build_semanal.py.

El Reports API no devuelve titulo ni marca; esos dos campos salen de Data
Kiosk (`groupByKey.productTitle` / `.brand`), que es la unica fuente que
los trae.
"""

import argparse
import json
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import settings
from spapi import data_kiosk, reports
from spapi.client import SpApiClient


ANCHO_COLUMNA = 31.2
HOJA = "Sheet0"

CAB_VENTAS = [
    "ASIN", "Título del Producto", "Marca", "Ganancia por envíos",
    "COGS por envíos", "Unidades enviadas", "Devoluciones del cliente",
    None, None, None,
]

CAB_INVENTARIO = [
    "ASIN", "Título del Producto", "Marca",
    "Porcentaje de confirmación del proveedor", "Recibido neto",
    "Unidades netas recibidas", "Cantidad de órdenes de compra abiertas",
    "Porcentaje de cumplimiento de recepción",
    "Tiempo total de entrega del proveedor (días)",
    "Inventario apto para la venta de más de 90 días",
    "Unidades aptas para la venta de más de 90 días",
    "Inventario apto para la venta disponible",
    "Unidades aptas para la venta disponibles",
    "Inventario no apto para la venta disponible",
    "Unidades no aptas para la venta disponibles",
]


def _dmy(d: date) -> str:
    return d.strftime("%d/%m/%y")


def metadatos(inicio: date, fin: date, periodo: str, n_cols: int) -> list:
    """Fila 1: los mismos filtros que estampa Vendor Central."""
    fila = [
        "Programa=[Retail]",
        "Vista del distribuidor=[Abastecimiento]",
        "Visto por=[ASIN]",
        "Países=[MX]",
        "Empresas=[ACER COMPUTEC MEXICO]",
        "Región=[es_MX]",
        "Moneda=[MXN]",
        f"Rango de generación de reportes=[{periodo}]",
        f"Rango de visualización=[{_dmy(inicio)} - {_dmy(fin)}]",
        f"Informe actualizado=[{_dmy(fin)}]\n",
    ]
    return fila + [""] * (n_cols - len(fila))


def _vacio(v):
    """Vendor Central deja la celda vacia donde la API devuelve null."""
    return "" if v is None else v


def _monto(v):
    return None if v is None else v.get("amount")


# --------------------------------------------------------------------------
# Catalogo (titulo + marca) — solo Data Kiosk los devuelve
# --------------------------------------------------------------------------

def catalogo(*jsonl_paths: Path) -> dict[str, tuple[str, str]]:
    cat: dict[str, tuple[str, str]] = {}
    for ruta in jsonl_paths:
        if not ruta or not Path(ruta).is_file():
            continue
        for linea in Path(ruta).read_text(encoding="utf-8").splitlines():
            if not linea.strip():
                continue
            for reg in json.loads(linea).get("metrics", []):
                k = reg.get("groupByKey", {})
                asin = k.get("asin")
                if asin and asin not in cat:
                    cat[asin] = (k.get("productTitle") or "", k.get("brand") or "")
    return cat


# --------------------------------------------------------------------------
# Construccion de las hojas
# --------------------------------------------------------------------------

def hoja_ventas(ws, filas_api: list[dict], cat: dict, inicio: date, fin: date) -> int:
    """Agrega el sell-out diario por ASIN y lo ordena por ganancia desc."""
    acc: dict[str, dict] = defaultdict(
        lambda: {"rev": 0.0, "cogs": 0.0, "units": 0, "ret": 0}
    )
    for r in filas_api:
        a = acc[r["asin"]]
        a["rev"] += (r.get("shippedRevenue") or {}).get("amount") or 0
        a["cogs"] += (r.get("shippedCogs") or {}).get("amount") or 0
        a["units"] += r.get("shippedUnits") or 0
        a["ret"] += r.get("customerReturns") or 0

    ws.append(metadatos(inicio, fin, "Semana", len(CAB_VENTAS)))
    ws.append(CAB_VENTAS)

    # Vendor Central omite los ASINs sin movimiento en la semana.
    vivos = {a: v for a, v in acc.items() if v["units"] or v["rev"] or v["ret"]}
    for asin, v in sorted(vivos.items(), key=lambda kv: -kv[1]["rev"]):
        titulo, marca = cat.get(asin, ("", ""))
        ws.append([
            asin, titulo, marca,
            round(v["rev"], 2), round(v["cogs"], 2), float(v["units"]),
            v["ret"] if v["ret"] else "",
            None, None, None,
        ])
    return len(vivos)


def hoja_inventario(ws, filas_api: list[dict], cat: dict, snap: date) -> int:
    """Snapshot de un dia, ordenado por ASIN descendente."""
    ws.append(metadatos(snap, snap, "Personalizado", len(CAB_INVENTARIO)))
    ws.append(CAB_INVENTARIO)

    for r in sorted(filas_api, key=lambda r: r["asin"], reverse=True):
        titulo, marca = cat.get(r["asin"], ("", ""))
        ws.append([
            r["asin"], titulo, marca,
            _vacio(r.get("vendorConfirmationRate")),
            _vacio(_monto(r.get("netReceivedInventoryCost"))),
            _vacio(r.get("netReceivedInventoryUnits")),
            _vacio(r.get("openPurchaseOrderUnits")),
            _vacio(r.get("receiveFillRate")),
            _vacio(r.get("averageVendorLeadTimeDays")),
            _vacio(_monto(r.get("aged90PlusDaysSellableInventoryCost"))),
            _vacio(r.get("aged90PlusDaysSellableInventoryUnits")),
            _vacio(_monto(r.get("sellableOnHandInventoryCost"))),
            _vacio(r.get("sellableOnHandInventoryUnits")),
            _vacio(_monto(r.get("unsellableOnHandInventoryCost"))),
            _vacio(r.get("unsellableOnHandInventoryUnits")),
        ])
    return len(filas_api)


def guardar(ws_builder, destino: Path, n_cols: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    ws_builder(ws)
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = ANCHO_COLUMNA
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)


# --------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(description="Exporta los xlsx de Vendor Central desde la API")
    p.add_argument("--descarga", type=date.fromisoformat,
                   help="el domingo en que se descargaria a mano; cubre "
                        "del domingo anterior al sabado previo (semana Amazon)")
    p.add_argument("--desde", type=date.fromisoformat)
    p.add_argument("--hasta", type=date.fromisoformat)
    p.add_argument("--destino", type=Path, default=settings.OUTPUT_DIR)
    p.add_argument("--sin-catalogo", action="store_true",
                   help="no consultar Data Kiosk; deja titulo y marca vacios")
    args = p.parse_args()

    if args.descarga:
        if args.descarga.weekday() != 6:
            p.error(f"--descarga debe ser un domingo; {args.descarga} es "
                    f"{args.descarga:%A}")
        fin = args.descarga - timedelta(days=1)     # sabado
        inicio = fin - timedelta(days=6)            # domingo anterior
    elif args.desde and args.hasta:
        inicio, fin = args.desde, args.hasta
    else:
        p.error("indica --descarga o --desde/--hasta")

    print(f"Rango: {inicio} ({inicio:%a}) -> {fin} ({fin:%a})")
    client = SpApiClient()

    crudo_v = reports.obtener_reporte(client, reports.SALES_REPORT, inicio, fin)
    if not crudo_v:
        print("Sin datos de ventas para el rango; no se genera nada.")
        return 1
    ventas = json.loads(crudo_v)["salesByAsin"]

    crudo_i = reports.obtener_reporte(client, reports.INVENTORY_REPORT, fin, fin)
    inventario = json.loads(crudo_i)["inventoryByAsin"] if crudo_i else []

    cat: dict[str, tuple[str, str]] = {}
    if not args.sin_catalogo:
        settings.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        rutas = []
        for etiqueta, contenido in (
            ("snap", data_kiosk.ejecutar(client, fin, fin)),
            ("ventas", data_kiosk.ejecutar(client, inicio, fin, solo_ventas=True)),
        ):
            if contenido:
                r = settings.OUTPUT_DIR / f"cat_{etiqueta}_{fin:%Y%m%d}.jsonl"
                r.write_text(contenido, encoding="utf-8")
                rutas.append(r)
        cat = catalogo(*rutas)
        print(f"  catalogo: {len(cat)} ASINs con titulo/marca")

    # Convencion de Vendor Central: el archivo lleva la fecha del domingo
    # en que se descarga, no la del ultimo dia de datos.
    sello = f"{fin + timedelta(days=1):%Y_%m_%d}"
    fv = args.destino / f"Amazon_SO_WTD_Ventas_{sello}.xlsx"
    fi = args.destino / f"Amazon_SO_WTD_Inventario_{sello}.xlsx"

    n_v = [0]
    guardar(lambda ws: n_v.append(hoja_ventas(ws, ventas, cat, inicio, fin)),
            fv, len(CAB_VENTAS))
    print(f"  -> {fv}  ({n_v[-1]} ASINs)")

    n_i = [0]
    guardar(lambda ws: n_i.append(hoja_inventario(ws, inventario, cat, fin)),
            fi, len(CAB_INVENTARIO))
    print(f"  -> {fi}  ({n_i[-1]} ASINs)")

    faltan = sum(1 for a in {r["asin"] for r in ventas} | {r["asin"] for r in inventario}
                 if a not in cat)
    if faltan:
        print(f"  AVISO: {faltan} ASINs sin titulo/marca (quedan vacios)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
